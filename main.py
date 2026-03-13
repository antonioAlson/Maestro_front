import customtkinter as ctk
from PIL import Image
import os
import subprocess
import threading
import re
import sys
import tkinter as tk

# Importar módulo update_dates
from scripts.data_update import update_dates


class SidebarApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Maestro")
        self.root.geometry("1000x500")
        self.set_window_icon()
        # Configurar tema
        ctk.set_appearance_mode("dark")  # "light" ou "dark"
        ctk.set_default_color_theme("blue")  # "blue", "green", "dark-blue"
        
        # Centralizar janela na tela
        self.center_window(1000, 500)
        
        # Estado da sidebar (True = aberta, False = fechada)
        self.sidebar_expanded = False
        
        # Larguras da sidebar
        self.sidebar_width_expanded = 200
        self.sidebar_width_collapsed = 60
        
        # Carregar imagens
        self.load_images()
        
        # Criar container principal
        self.main_container = ctk.CTkFrame(self.root, fg_color="transparent")
        self.main_container.pack(fill="both", expand=True)
        
        # Criar sidebar
        self.create_sidebar()
        
        # Variável para rastrear o botão de menu ativo
        self.active_menu_button = None
        
        # Se a sidebar começa colapsada, esconder textos dos botões
        if not self.sidebar_expanded:
            self.update_button_texts("")
        
        # Criar área de conteúdo
        self.create_content_area()
        
        # Abrir tela Home por padrão
        self.home_action()

    def set_window_icon(self):
        """Define o ícone da janela principal"""
        icon_ico_path = os.path.join(os.path.dirname(__file__), "img", "icone.ico")
        
        if not os.path.exists(icon_ico_path):
            print(f"Ícone não encontrado: {icon_ico_path}")
            return
        
        try:
            # Aplicar ícone .ico no Windows
            self.root.iconbitmap(icon_ico_path)
            # Reaplicar após inicialização para garantir persistência
            self.root.after(100, lambda: self.root.iconbitmap(icon_ico_path))
            print(f"Ícone aplicado: {icon_ico_path}")
        except Exception as e:
            print(f"Erro ao definir ícone da janela: {e}")
    
    def center_window(self, width, height):
        """Centraliza a janela na tela"""
        # Obter dimensões da tela
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # Calcular posição x e y para centralizar
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        
        # Definir geometria com posição
        self.root.geometry(f"{width}x{height}+{x}+{y}")

    def load_images(self):
        """Carrega as imagens dos ícones"""
        self.images = {}
        img_dir = "img"
        icon_size = (24, 24)
        
        # Mapeamento de ícones
        image_files = {
            "home": "home.png",
            "dashboard": "dashboard.png",
            "pcp": "pcp.png",
            "settings": "settings.png",
            "exit": "exit.png"
        }
        
        for key, filename in image_files.items():
            img_path = os.path.join(img_dir, filename)
            if os.path.exists(img_path):
                try:
                    img = Image.open(img_path)
                    self.images[key] = ctk.CTkImage(light_image=img, dark_image=img, size=icon_size)
                except Exception as e:
                    print(f"Erro ao carregar imagem {filename}: {e}")
                    self.images[key] = None
            else:
                print(f"Imagem não encontrada: {img_path}")
                self.images[key] = None
        
    def create_sidebar(self):
        """Cria a sidebar com menu"""
        initial_width = self.sidebar_width_expanded if self.sidebar_expanded else self.sidebar_width_collapsed
        self.sidebar = ctk.CTkFrame(
            self.main_container,
            width=initial_width,
            corner_radius=0,
        )
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)
        
        # Botão de toggle no topo
        self.toggle_btn = ctk.CTkButton(
            self.sidebar,
            text="☰",
            command=self.toggle_sidebar,
            width=40,
            height=40,
            font=ctk.CTkFont(size=20, weight="bold"),
            corner_radius=8,
            fg_color="transparent",
            hover_color="#3a3a5e"
        )
        self.toggle_btn.pack(pady=15, padx=10, anchor="w")
        
        # Container para os botões do menu (parte superior)
        self.menu_container = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        self.menu_container.pack(fill="both", expand=True, pady=10)
        
        # Itens do menu principal (imagem_key, texto, comando)
        menu_items = [
            ("home", "Home", self.home_action),
            ("dashboard", "Dashboard", self.dashboard_action),
            ("pcp", "PCP", self.pcp_action),
            ("settings", "Configurações", self.settings_action),
        ]
        
        # Criar botões do menu
        self.menu_buttons = []
        for img_key, text, command in menu_items:
            btn = self.create_menu_button(self.menu_container, img_key, text, command)
            self.menu_buttons.append((btn, img_key, text))
        
        # Container para o botão Exit (parte inferior)
        self.exit_container = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        self.exit_container.pack(side="bottom", fill="x", pady=10)
        
        # Criar botão Exit
        self.exit_btn = self.create_menu_button(self.exit_container, "exit", "Sair", self.exit_action)
    

    def create_menu_button(self, container, img_key, text, command):
        """Cria um botão do menu"""
        img = self.images.get(img_key)
        
        # Frame wrapper para controlar overflow
        btn_wrapper = ctk.CTkFrame(container, fg_color="transparent")
        btn_wrapper.pack(fill="x", pady=3, padx=8)
        
        btn = ctk.CTkButton(
            btn_wrapper,
            text=f"  {text}",
            image=img,
            command=command,
            anchor="w",
            height=40,
            corner_radius=8,
            font=ctk.CTkFont(size=13),
            compound="left",
            fg_color="transparent",
            hover_color="#3a3a5e"
        )
        btn.pack(fill="x")
        
        return btn
    
    def toggle_sidebar(self):
        """Alterna entre sidebar aberta e fechada"""
        if self.sidebar_expanded:
            # Fechar sidebar
            self.sidebar_expanded = False
            self.sidebar.configure(width=self.sidebar_width_collapsed)
            self.update_button_texts("")
            self.toggle_btn.configure(text="☰")
        else:
            # Abrir sidebar
            self.sidebar_expanded = True
            self.sidebar.configure(width=self.sidebar_width_expanded)
            self.root.after(50, lambda: self.update_button_texts("show"))
            self.toggle_btn.configure(text="✕")
    
    def update_button_texts(self, mode):
        """Atualiza os textos dos botões"""
        if mode == "":
            # Esconder textos
            for btn, img_key, text in self.menu_buttons:
                btn.configure(text="")
            self.exit_btn.configure(text="")
        else:
            # Mostrar textos
            for btn, img_key, text in self.menu_buttons:
                btn.configure(text=f"  {text}")
            self.exit_btn.configure(text="  Sair")
    
    def set_active_menu_button(self, button_name):
        """Marca um botão do menu como ativo"""
        for btn, img_key, text in self.menu_buttons:
            if img_key == button_name:
                # Botão ativo - cor de destaque
                btn.configure(fg_color="#3a3a5e")
                self.active_menu_button = img_key
            else:
                # Botões inativos - transparente
                btn.configure(fg_color="transparent")

    def create_content_area(self):
        """Cria a área de conteúdo principal"""
        self.content_area = ctk.CTkFrame(self.main_container, corner_radius=0, fg_color="#3d3d3d")
        self.content_area.pack(side="left", fill="both", expand=True)
        
        # Frame para exibir conteúdo dinâmico
        self.dynamic_content = ctk.CTkFrame(self.content_area, fg_color="transparent")
        self.dynamic_content.pack(fill="both", expand=True, padx=10, pady=10)
    
    def clear_content(self):
        """Limpa o conteúdo dinâmico"""
        for widget in self.dynamic_content.winfo_children():
            widget.destroy()
    
    def show_content(self, title, description, title_anchor="center", title_img=None):
        """Mostra conteúdo na área principal"""
        self.clear_content()
        
        # Card de conteúdo
        card = ctk.CTkFrame(self.dynamic_content, corner_radius=10)
        card.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Título do card com imagem
        # Adicionar espaçamento entre imagem e texto
        title_text = f"  {title}" if title_img else title
        title_label = ctk.CTkLabel(
            card,
            text=title_text,
            font=ctk.CTkFont(size=24, weight="bold"),
            image=title_img,
            compound="left"
        )
        title_label.pack(pady=30, anchor=title_anchor, padx=30)
    
    # Ações dos botões do menu
    def home_action(self):
        self.set_active_menu_button("home")
        self.clear_content()
        
        # Card de conteúdo
        card = ctk.CTkFrame(self.dynamic_content, corner_radius=10)
        card.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Container centralizado para o título
        title_container = ctk.CTkFrame(card, fg_color="transparent")
        title_container.place(relx=0.5, rely=0.5, anchor="center")
        
        # Frame para título com gradiente simulado
        title_frame = ctk.CTkFrame(title_container, fg_color="transparent")
        title_frame.pack()
        
        # Título MAESTRO estilizado com efeito gradiente (letras com cores diferentes)
        letters_colors = [
            ("#1f6aa5", "M"),
            ("#2674ae", "A"),
            ("#2d7eb7", "E"),
            ("#3488c0", "S"),
            ("#3b92c9", "T"),
            ("#429cd2", "R"),
            ("#49a6db", "O")
        ]
        
        for color, letter in letters_colors:
            letter_label = ctk.CTkLabel(
                title_frame,
                text=letter,
                font=ctk.CTkFont(family="Segoe UI", size=80, weight="bold"),
                text_color=color
            )
            letter_label.pack(side="left", padx=1)
        
        # Subtítulo
        subtitle = ctk.CTkLabel(
            title_container,
            text="Sistema de Gestão PCP/PROCESSOS",
            font=ctk.CTkFont(family="Segoe UI", size=20),
            text_color="gray60"
        )
        subtitle.pack(pady=(15, 0))
    
    def dashboard_action(self):
        self.set_active_menu_button("dashboard")
        self.show_content(
            "Dashboard",
            "Dashboard com métricas e estatísticas importantes do sistema.",
            title_anchor="w",
            title_img=self.images.get("dashboard")
        )
    
    def pcp_action(self):
        """Mostra conteúdo PCP com layout organizado e ações em grade"""
        self.set_active_menu_button("pcp")
        self.clear_content()
        
        # Card principal da tela PCP
        card = ctk.CTkFrame(self.dynamic_content, corner_radius=12, fg_color="#2b2b2b")
        card.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Cabecalho da area PCP
        header_frame = ctk.CTkFrame(card, fg_color="transparent")
        header_frame.pack(fill="x", padx=26, pady=(22, 6))

        title_text = "  PCP"
        title_label = ctk.CTkLabel(
            header_frame,
            text=title_text,
            font=ctk.CTkFont(size=26, weight="bold"),
            image=self.images.get("pcp"),
            compound="left"
        )
        title_label.pack(anchor="w")

        subtitle_label = ctk.CTkLabel(
            header_frame,
            text="Selecione uma rotina para iniciar o processamento.",
            font=ctk.CTkFont(size=13),
            text_color="gray75"
        )
        subtitle_label.pack(anchor="w", pady=(4, 0))

        divider = ctk.CTkFrame(card, height=1, corner_radius=0, fg_color="#3a3a3a")
        divider.pack(fill="x", padx=26, pady=(8, 10))
        
        # Container de acoes
        actions_frame = ctk.CTkFrame(card, corner_radius=12, fg_color="#323232")
        actions_frame.pack(fill="both", expand=True, padx=26, pady=(0, 22))

        actions_title = ctk.CTkLabel(
            actions_frame,
            text="Ações PCP",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="gray85"
        )
        actions_title.pack(anchor="w", padx=18, pady=(14, 4))

        buttons_container = ctk.CTkFrame(actions_frame, fg_color="transparent")
        buttons_container.pack(fill="x", padx=14, pady=(0, 14))

        # Grid 2x2 com botoes uniformes e altura controlada
        buttons_container.grid_columnconfigure(0, weight=1, uniform="pcp_col")
        buttons_container.grid_columnconfigure(1, weight=1, uniform="pcp_col")
        buttons_container.grid_rowconfigure((0, 1), weight=0)
        
        # Definir botoes
        button_names = [
            "Gerar Relatório",
            "Adicionar Datas",
            "Reprogramar Atrasos",
            "Imprimir OS"
        ]
        
        # Criar botoes em grade 2x2 (2 colunas, 2 linhas)
        for i, btn_name in enumerate(button_names):
            row = i // 2
            col = i % 2
            btn = ctk.CTkButton(
                buttons_container,
                text=btn_name,
                height=42,
                font=ctk.CTkFont(size=13, weight="bold"),
                corner_radius=8,
                fg_color="#1f6aa5",
                hover_color="#2f7dc2",
                command=lambda name=btn_name: self.pcp_routine_action(name)
            )
            btn.grid(row=row, column=col, padx=6, pady=6, sticky="ew")
    
    def show_loading_popup(self, message="Processando..."):
        """Mostra popup de carregamento centralizado"""
        self.loading_popup = ctk.CTkToplevel(self.root)
        self.loading_popup.title("Aguarde")
        self.loading_popup.geometry("550x400")
        self.loading_popup.resizable(False, False)
        
        # Centralizar baseado na janela principal
        self.loading_popup.transient(self.root)
        self.loading_popup.grab_set()
        
        # Atualizar para garantir que as dimensões estejam corretas
        self.loading_popup.update_idletasks()
        
        # Calcular posição central
        root_x = self.root.winfo_x()
        root_y = self.root.winfo_y()
        root_width = self.root.winfo_width()
        root_height = self.root.winfo_height()
        
        popup_width = 550
        popup_height = 400
        
        x = root_x + (root_width - popup_width) // 2
        y = root_y + (root_height - popup_height) // 2
        
        self.loading_popup.geometry(f"{popup_width}x{popup_height}+{x}+{y}")
        
        # Mensagem
        label = ctk.CTkLabel(
            self.loading_popup,
            text=message,
            font=ctk.CTkFont(size=16, weight="bold")
        )
        label.pack(pady=20)
        
        # Progress bar determinada
        self.progress_bar = ctk.CTkProgressBar(self.loading_popup, mode="determinate")
        self.progress_bar.pack(pady=10, padx=40, fill="x")
        self.progress_bar.set(0)
        
        # Label de porcentagem
        self.progress_label = ctk.CTkLabel(
            self.loading_popup,
            text="0%",
            font=ctk.CTkFont(size=14)
        )
        self.progress_label.pack(pady=5)
        
        # Textbox para mostrar log em tempo real
        self.log_textbox = ctk.CTkTextbox(
            self.loading_popup,
            height=200,
            font=ctk.CTkFont(size=11)
        )
        self.log_textbox.pack(pady=(10, 20), padx=20, fill="both", expand=True)
        
        return self.loading_popup
    
    def update_progress(self, value):
        """Atualiza a barra de progresso e o label"""
        if hasattr(self, 'progress_bar') and self.progress_bar:
            self.progress_bar.set(value)
        if hasattr(self, 'progress_label') and self.progress_label:
            self.progress_label.configure(text=f"{int(value * 100)}%")
    
    def append_log(self, text):
        """Adiciona texto ao log do popup de carregamento"""
        if hasattr(self, 'log_textbox') and self.log_textbox:
            self.log_textbox.insert("end", text + "\n")
            self.log_textbox.see("end")  # Auto-scroll para o final
    
    def add_close_button_to_popup(self):
        """Adiciona botões de ação ao popup de carregamento após conclusão"""
        if hasattr(self, 'loading_popup') and self.loading_popup:
            # Frame para os botões
            buttons_frame = ctk.CTkFrame(self.loading_popup, fg_color="transparent")
            buttons_frame.pack(pady=(0, 10))
            
            # Botão Abrir Pasta
            open_folder_btn = ctk.CTkButton(
                buttons_frame,
                text="📁 Abrir Pasta",
                command=self.open_download_folder,
                width=140,
                height=35,
                font=ctk.CTkFont(size=13, weight="bold"),
                fg_color="#1f6aa5",
                hover_color="#2f7dc2"
            )
            open_folder_btn.pack(side="left", padx=5)
            
            # Botão Fechar
            close_btn = ctk.CTkButton(
                buttons_frame,
                text="Fechar",
                command=self.close_loading_popup,
                width=140,
                height=35,
                font=ctk.CTkFont(size=13, weight="bold"),
                fg_color="#2a9d2a",
                hover_color="#238a23"
            )
            close_btn.pack(side="left", padx=5)
    
    def open_download_folder(self):
        """Abre a pasta onde o ZIP foi salvo"""
        folder_path = os.path.join("src", "temp", "download_os")
        abs_path = os.path.abspath(folder_path)
        
        # Criar pasta se não existir
        os.makedirs(abs_path, exist_ok=True)
        
        # Abrir no Windows Explorer
        if os.path.exists(abs_path):
            os.startfile(abs_path)
    
    def show_success_message(self, script_name, file_path=None):
        """Mostra mensagem de sucesso no popup"""
        if hasattr(self, 'loading_popup') and self.loading_popup:
            # Ajustar altura do popup para caber todos os elementos
            self.loading_popup.geometry("350x220")
            
            # Limpar widgets existentes
            for widget in self.loading_popup.winfo_children():
                widget.destroy()
            
            # Mensagem de sucesso
            success_label = ctk.CTkLabel(
                self.loading_popup,
                text="✓ Concluído com sucesso!",
                font=ctk.CTkFont(size=18, weight="bold"),
                text_color="#90EE90"
            )
            success_label.pack(pady=25)
            
            # Detalhes
            detail_label = ctk.CTkLabel(
                self.loading_popup,
                text=f"{script_name} executado com sucesso.",
                font=ctk.CTkFont(size=13)
            )
            detail_label.pack(pady=5)
            
            # Container para botões
            buttons_frame = ctk.CTkFrame(self.loading_popup, fg_color="transparent")
            buttons_frame.pack(pady=20)
            
            # Botão abrir Excel (se houver caminho e for arquivo Excel)
            if file_path and file_path.endswith('.xlsx'):
                view_btn = ctk.CTkButton(
                    buttons_frame,
                    text="Abrir Excel",
                    command=lambda: [self.open_file(file_path), self.close_loading_popup()],
                    width=130,
                    height=35,
                    font=ctk.CTkFont(size=13, weight="bold")
                )
                view_btn.pack(side="left", padx=5)
            
            # Botão fechar
            close_btn = ctk.CTkButton(
                buttons_frame,
                text="Fechar",
                command=self.close_loading_popup,
                width=120,
                height=35,
                font=ctk.CTkFont(size=13, weight="bold")
            )
            close_btn.pack(side="left", padx=5)
    
    def open_file(self, file_path):
        """Abre o arquivo com o aplicativo padrão do sistema"""
        try:
            if os.path.exists(file_path):
                os.startfile(file_path)  # Windows
                print(f"Abrindo arquivo: {file_path}")
            else:
                print(f"Arquivo não encontrado: {file_path}")
        except Exception as e:
            print(f"Erro ao abrir arquivo: {e}")
    
    def save_and_update_jira(self, file_path, entries_data):
        """Salva alterações no Excel e atualiza no Jira"""
        # Primeiro salvar no Excel (sem mostrar confirmação)
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from datetime import datetime
            
            # Ler arquivo Excel
            df = pd.read_excel(file_path)
            
            # Converter a coluna de data para datetime se necessário
            if "Prev. de Entrega" in df.columns:
                df["Prev. de Entrega"] = pd.to_datetime(df["Prev. de Entrega"], errors='coerce')
            
            # Atualizar todos os valores
            for row_idx, column, entry in entries_data:
                new_value = entry.get().strip()
                
                if column == "Prev. de Entrega":
                    if new_value:
                        try:
                            if "/" in new_value:
                                date_obj = datetime.strptime(new_value, "%d/%m/%Y")
                            elif "-" in new_value and len(new_value) == 10:
                                try:
                                    date_obj = datetime.strptime(new_value, "%Y-%m-%d")
                                except:
                                    date_obj = datetime.strptime(new_value, "%d-%m-%Y")
                            else:
                                date_obj = pd.to_datetime(new_value)
                            
                            df.at[row_idx, column] = pd.Timestamp(date_obj)
                        except (ValueError, Exception) as e:
                            print(f"Aviso: Formato de data inválido '{new_value}' na linha {row_idx}")
                            df.at[row_idx, column] = pd.NaT
                    else:
                        df.at[row_idx, column] = pd.NaT
                else:
                    df.at[row_idx, column] = new_value if new_value else ""
            
            # Salvar arquivo
            df.to_excel(file_path, index=False)
            
            # Reaplicar hyperlinks se a coluna "Chave" existir
            if "Chave" in df.columns:
                wb = load_workbook(file_path)
                ws = wb.active
                
                for idx, link in enumerate(df["Chave"], start=2):
                    cell = ws[f'C{idx}']
                    cell.hyperlink = link
                    cell.style = "Hyperlink"
                
                if 'F' in ws.column_dimensions:
                    ws.column_dimensions['F'].width = 11
                
                wb.save(file_path)
            
            print(f"✓ Dados salvos no Excel!")
            
            # Aguardar e atualizar Jira
            self.root.after(500, lambda: self.update_jira_dates(file_path))
            
        except Exception as e:
            print(f"Erro ao salvar alterações: {e}")
            import traceback
            traceback.print_exc()
    
    def save_all_excel_changes(self, file_path, entries_data):
        """Salva todas as alterações dos campos de entrada no Excel"""
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from datetime import datetime
            
            # Ler arquivo Excel
            df = pd.read_excel(file_path)
            
            # Converter a coluna de data para datetime se necessário
            if "Prev. de Entrega" in df.columns:
                # Converter para datetime, erros viram NaT
                df["Prev. de Entrega"] = pd.to_datetime(df["Prev. de Entrega"], errors='coerce')
            
            # Atualizar todos os valores
            for row_idx, column, entry in entries_data:
                new_value = entry.get().strip()
                
                # Se for a coluna de data
                if column == "Prev. de Entrega":
                    if new_value:
                        try:
                            # Tentar converter de dd/mm/yyyy para datetime
                            if "/" in new_value:
                                date_obj = datetime.strptime(new_value, "%d/%m/%Y")
                            elif "-" in new_value and len(new_value) == 10:
                                # Se estiver no formato YYYY-MM-DD ou DD-MM-YYYY
                                try:
                                    date_obj = datetime.strptime(new_value, "%Y-%m-%d")
                                except:
                                    date_obj = datetime.strptime(new_value, "%d-%m-%Y")
                            else:
                                # Tentar outros formatos
                                date_obj = pd.to_datetime(new_value)
                            
                            # Atualizar com Timestamp do pandas
                            df.at[row_idx, column] = pd.Timestamp(date_obj)
                        except (ValueError, Exception) as e:
                            # Se não conseguir converter, deixar NaT
                            print(f"Aviso: Formato de data inválido '{new_value}' na linha {row_idx}")
                            df.at[row_idx, column] = pd.NaT
                    else:
                        # Se estiver vazio, usar NaT
                        df.at[row_idx, column] = pd.NaT
                else:
                    # Para outras colunas
                    df.at[row_idx, column] = new_value if new_value else ""
            
            # Salvar arquivo
            df.to_excel(file_path, index=False)
            
            # Reaplicar hyperlinks se a coluna "Chave" existir
            if "Chave" in df.columns:
                wb = load_workbook(file_path)
                ws = wb.active
                
                for idx, link in enumerate(df["Chave"], start=2):
                    cell = ws[f'C{idx}']
                    cell.hyperlink = link
                    cell.style = "Hyperlink"
                
                # Ajustar largura da coluna SITUAÇÃO (coluna F) para 80px
                if 'F' in ws.column_dimensions:
                    ws.column_dimensions['F'].width = 11
                
                wb.save(file_path)
            
            print(f"✓ Todas as alterações foram salvas com sucesso!")
            
            # Mostrar mensagem de confirmação visual
            self.show_save_confirmation()
            
        except Exception as e:
            print(f"Erro ao salvar alterações: {e}")
            import traceback
            traceback.print_exc()
    
    def update_jira_dates(self, file_path):
        """Atualiza as datas no Jira baseado no arquivo Excel"""
        try:
            import pandas as pd
            import requests
            from requests.auth import HTTPBasicAuth
            from datetime import datetime
            from dotenv import load_dotenv
            import os
            
            # Carregar variáveis de ambiente
            load_dotenv()
            
            JIRA_URL = os.getenv("JIRA_URL")
            EMAIL = os.getenv("EMAIL")
            API_TOKEN = os.getenv("API_TOKEN")
            CAMPO_PREVISAO = "customfield_10245"
            
            # Ler arquivo Excel
            df = pd.read_excel(file_path)
            
            # Filtrar apenas linhas com ID e data preenchidos
            df_update = df[df["ID"].notna() & df["Prev. de Entrega"].notna()].copy()
            
            if len(df_update) == 0:
                print("Nenhuma data para atualizar no Jira")
                return
            
            print(f"Atualizando {len(df_update)} issues no Jira...")
            
            headers = {
                "Accept": "application/json",
                "Content-Type": "application/json"
            }
            
            auth = HTTPBasicAuth(EMAIL, API_TOKEN)
            
            success_count = 0
            error_count = 0
            
            for index, row in df_update.iterrows():
                issue_id = row["ID"]
                data_entrega = row["Prev. de Entrega"]
                
                # Converter para formato YYYY-MM-DD
                try:
                    if isinstance(data_entrega, pd.Timestamp):
                        data_formatada = data_entrega.strftime("%Y-%m-%d")
                    elif isinstance(data_entrega, str):
                        # Tentar converter string para datetime
                        if "/" in data_entrega:
                            date_obj = datetime.strptime(data_entrega, "%d/%m/%Y")
                        else:
                            date_obj = datetime.strptime(data_entrega, "%Y-%m-%d")
                        data_formatada = date_obj.strftime("%Y-%m-%d")
                    else:
                        print(f"Formato não reconhecido para {issue_id}: {data_entrega}")
                        error_count += 1
                        continue
                    
                    # Fazer requisição ao Jira
                    url = f"{JIRA_URL}/rest/api/3/issue/{issue_id}"
                    payload = {
                        "fields": {
                            CAMPO_PREVISAO: data_formatada
                        }
                    }
                    
                    response = requests.put(url, headers=headers, auth=auth, json=payload)
                    
                    if response.status_code == 204:
                        print(f"✓ {issue_id} atualizado para {data_formatada}")
                        success_count += 1
                    else:
                        print(f"✗ Erro ao atualizar {issue_id}: {response.status_code} - {response.text}")
                        error_count += 1
                        
                except Exception as e:
                    print(f"✗ Erro ao processar {issue_id}: {e}")
                    error_count += 1
            
            # Mostrar resultado
            print(f"\nAtualização concluída: {success_count} sucesso, {error_count} erros")
            self.show_jira_update_result(success_count, error_count)
            
        except Exception as e:
            print(f"Erro ao atualizar Jira: {e}")
            import traceback
            traceback.print_exc()
    
    def show_jira_update_result(self, success_count, error_count):
        """Mostra resultado da atualização do Jira"""
        result_popup = ctk.CTkToplevel(self.root)
        result_popup.title("Resultado da Atualização")
        result_popup.geometry("350x180")
        result_popup.resizable(False, False)
        
        # Centralizar
        result_popup.transient(self.root)
        result_popup.update_idletasks()
        
        root_x = self.root.winfo_x()
        root_y = self.root.winfo_y()
        root_width = self.root.winfo_width()
        root_height = self.root.winfo_height()
        
        x = root_x + (root_width - 350) // 2
        y = root_y + (root_height - 180) // 2
        
        result_popup.geometry(f"350x180+{x}+{y}")
        
        # Mensagem
        title = "✓ Atualização Concluída!" if error_count == 0 else "⚠️ Atualização Parcial"
        color = "#90EE90" if error_count == 0 else "orange"
        
        label = ctk.CTkLabel(
            result_popup,
            text=title,
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=color
        )
        label.pack(pady=20)
        
        # Detalhes
        detail_text = f"✓ Sucesso: {success_count}\n✗ Erros: {error_count}"
        detail_label = ctk.CTkLabel(
            result_popup,
            text=detail_text,
            font=ctk.CTkFont(size=14)
        )
        detail_label.pack(pady=10)
        
        # Botão fechar
        close_btn = ctk.CTkButton(
            result_popup,
            text="Fechar",
            command=result_popup.destroy,
            width=120,
            height=35,
            font=ctk.CTkFont(size=13, weight="bold")
        )
        close_btn.pack(pady=15)
    
    def show_save_confirmation(self):
        """Mostra mensagem temporária de confirmação de salvamento"""
        # Criar popup temporário
        confirm_popup = ctk.CTkToplevel(self.root)
        confirm_popup.title("Sucesso")
        confirm_popup.geometry("300x120")
        confirm_popup.resizable(False, False)
        
        # Centralizar
        confirm_popup.transient(self.root)
        confirm_popup.update_idletasks()
        
        root_x = self.root.winfo_x()
        root_y = self.root.winfo_y()
        root_width = self.root.winfo_width()
        root_height = self.root.winfo_height()
        
        x = root_x + (root_width - 300) // 2
        y = root_y + (root_height - 120) // 2
        
        confirm_popup.geometry(f"300x120+{x}+{y}")
        
        # Mensagem
        label = ctk.CTkLabel(
            confirm_popup,
            text="✓ Dados salvos com sucesso!",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#90EE90"
        )
        label.pack(pady=30)
        
        # Fechar automaticamente após 1.5 segundos
        self.root.after(1500, confirm_popup.destroy)
    
    def view_excel_data(self, file_path):
        """Visualiza os dados do Excel em uma nova janela"""
        try:
            print(f"Iniciando visualização de dados: {file_path}")
            
            # Fechar popup de sucesso
            self.close_loading_popup()
            
            # Ler arquivo Excel
            import pandas as pd
            print(f"Lendo arquivo Excel...")
            df = pd.read_excel(file_path)
            print(f"Arquivo lido com sucesso. {len(df)} linhas encontradas")
            
            # Criar cópia para visualização
            df_visual = df.copy()
            
            # Lista para armazenar referências aos campos editáveis
            entries_data = []
            
            # Combinar Tipo de issue e Resumo em uma única coluna
            if "Tipo de issue" in df_visual.columns and "Resumo" in df_visual.columns:
                # Converter para string para evitar erros de concatenação
                df_visual["Tipo - OS"] = df_visual["Tipo de issue"].astype(str) + " - " + df_visual["Resumo"].astype(str)
                # Remover as colunas originais na visualização
                df_visual = df_visual.drop(columns=["Tipo de issue", "Resumo"])
            elif "Issue" in df_visual.columns and "Tipo - OS" not in df_visual.columns:
                df_visual = df_visual.rename(columns={"Issue": "Tipo - OS"})
            elif "OS" in df_visual.columns and "Tipo - OS" not in df_visual.columns:
                df_visual = df_visual.rename(columns={"OS": "Tipo - OS"})
            
            # Filtrar colunas (remover ID e Chave)
            columns_to_hide = ["ID", "Chave"]
            display_columns = [col for col in df_visual.columns if col not in columns_to_hide]
            
            # Reorganizar para colocar Tipo - OS em primeiro
            if "Tipo - OS" in display_columns:
                display_columns.remove("Tipo - OS")
                display_columns.insert(0, "Tipo - OS")
            
            print(f"Colunas a exibir: {display_columns}")
            print("Criando nova janela...")
            
            # Criar nova janela
            excel_window = ctk.CTkToplevel(self.root)
            excel_window.title("Visualização de Dados - " + os.path.basename(file_path))
            excel_window.geometry("900x500")
            print("Janela criada")
            
            # Centralizar a nova janela
            excel_window.update_idletasks()
            width = 900
            height = 500
            screen_width = excel_window.winfo_screenwidth()
            screen_height = excel_window.winfo_screenheight()
            x = (screen_width - width) // 2
            y = (screen_height - height) // 2
            excel_window.geometry(f"{width}x{height}+{x}+{y}")
            print(f"Janela posicionada em {x},{y}")
            
            # Trazer janela para frente
            excel_window.lift()
            excel_window.focus_force()
            excel_window.attributes('-topmost', True)
            excel_window.after(100, lambda: excel_window.attributes('-topmost', False))
            
            # Card de conteúdo
            card = ctk.CTkFrame(excel_window, corner_radius=10)
            card.pack(fill="both", expand=True, padx=10, pady=10)
            print("Card criado")
            
            # Título
            title_label = ctk.CTkLabel(
                card,
                text="  Visualização de Dados",
                font=ctk.CTkFont(size=18, weight="bold")
            )
            title_label.pack(pady=10, anchor="w", padx=15)
            print("Título adicionado")
            
            # Info do arquivo
            info_label = ctk.CTkLabel(
                card,
                text=f"Arquivo: {os.path.basename(file_path)} | Total de registros: {len(df_visual)}",
                font=ctk.CTkFont(size=10),
                text_color="gray70"
            )
            info_label.pack(pady=3, anchor="w", padx=15)
            print("Info adicionada")
            
            # Frame scrollable para a tabela
            table_frame = ctk.CTkScrollableFrame(card, height=280)
            table_frame.pack(fill="both", expand=True, padx=15, pady=10)
            print("Frame scrollable criado")

            def largura_coluna(column_name):
                if column_name in ["Tipo - OS", "Status", "SITUAÇÃO"]:
                    return 50
                if column_name == "Veículo":
                    return 310
                if column_name == "Prev. de Entrega":
                    return 80
                return 150

            col_widths = [largura_coluna(col) for col in display_columns]
            for col_idx, col_width in enumerate(col_widths):
                table_frame.grid_columnconfigure(col_idx, weight=0, minsize=col_width)
            table_frame.grid_rowconfigure(0, minsize=38)
            
            # Criar cabeçalho da tabela
            print(f"Criando cabeçalho com {len(display_columns)} colunas")
            for col_idx, column in enumerate(display_columns):
                header_frame = ctk.CTkFrame(
                    table_frame,
                    fg_color="gray25",
                    border_width=1,
                    border_color="gray40",
                    width=col_widths[col_idx],
                    height=38
                )
                header_frame.grid(row=0, column=col_idx, padx=1, pady=1, sticky="nsew")
                header_frame.grid_propagate(False)
                
                header = ctk.CTkLabel(
                    header_frame,
                    text=str(column),
                    font=ctk.CTkFont(size=12, weight="bold"),
                    anchor="w"
                )
                header.pack(padx=8, pady=6, fill="both", expand=True)
            print("Cabeçalho criado")
            
            # Preencher dados da tabela (limitado a 100 linhas para performance)
            max_rows = min(len(df_visual), 100)
            print(f"Preenchendo {max_rows} linhas de dados...")
            for row_idx in range(max_rows):
                table_frame.grid_rowconfigure(row_idx + 1, minsize=34)
                for col_idx, column in enumerate(display_columns):
                    value = df_visual[column].iloc[row_idx]
                    # Converter para string e limitar tamanho
                    cell_text = str(value) if pd.notna(value) else ""
                    if column == "Tipo - OS" and len(cell_text) > 20:
                        cell_text = cell_text[:17] + "..."
                    elif column == "Veículo" and len(cell_text) > 50:
                        cell_text = cell_text[:47] + "..."
                    elif len(cell_text) > 50 and column != "Prev. de Entrega":
                        cell_text = cell_text[:47] + "..."
                    
                    # Frame para criar borda da célula
                    cell_frame = ctk.CTkFrame(
                        table_frame,
                        fg_color="gray20",
                        border_width=1,
                        border_color="gray30",
                        width=col_widths[col_idx],
                        height=34
                    )
                    cell_frame.grid(row=row_idx + 1, column=col_idx, padx=1, pady=1, sticky="nsew")
                    cell_frame.grid_propagate(False)
                    
                    # Se for a coluna de data, criar um campo de entrada editável
                    if column == "Prev. de Entrega":
                        entry = ctk.CTkEntry(
                            cell_frame,
                            font=ctk.CTkFont(size=11),
                            width=max(60, col_widths[col_idx] - 10),
                            height=28,
                            justify="center"
                        )
                        entry.insert(0, cell_text)
                        entry.pack(padx=4, pady=2, fill="both", expand=True)
                        
                        # Armazenar referência para salvar depois
                        entries_data.append((row_idx, column, entry))
                    else:
                        # Para outras colunas, usar label como antes
                        cell = ctk.CTkLabel(
                            cell_frame, 
                            text=cell_text,
                            font=ctk.CTkFont(size=11),
                            anchor="w"
                        )
                        cell.pack(padx=8, pady=4, fill="both", expand=True)
            
            print(f"Tabela preenchida com {max_rows} linhas e {len(entries_data)} campos editáveis")
            
            # Aviso se houver mais linhas
            if len(df) > 100:
                warning_label = ctk.CTkLabel(
                    card,
                    text=f"⚠️ Exibindo apenas as primeiras 100 linhas de {len(df)}",
                    font=ctk.CTkFont(size=9),
                    text_color="orange"
                )
                warning_label.pack(pady=3, anchor="w", padx=15)
            
            # Container para botões
            buttons_container = ctk.CTkFrame(card, fg_color="transparent")
            buttons_container.pack(pady=8)
            print("Container de botões criado")
            
            # Botão Salvar e Atualizar Jira
            save_btn = ctk.CTkButton(
                buttons_container,
                text="💾 Salvar e Atualizar Jira",
                command=lambda: self.save_and_update_jira(file_path, entries_data),
                width=170,
                height=30,
                font=ctk.CTkFont(size=11, weight="bold"),
                fg_color="#2a9d2a",
                hover_color="#238a23"
            )
            save_btn.pack(side="left", padx=3)
            
            # Botão para abrir no Excel
            open_excel_btn = ctk.CTkButton(
                buttons_container,
                text="📄 Abrir Excel",
                command=lambda: self.open_file(file_path),
                width=120,
                height=30,
                font=ctk.CTkFont(size=11, weight="bold")
            )
            open_excel_btn.pack(side="left", padx=3)
            
            # Botão Fechar
            close_btn = ctk.CTkButton(
                buttons_container,
                text="✕ Fechar",
                command=excel_window.destroy,
                width=100,
                height=30,
                font=ctk.CTkFont(size=11, weight="bold"),
                fg_color="gray40",
                hover_color="gray50"
            )
            close_btn.pack(side="left", padx=3)
            
            print("Janela de visualização criada com sucesso!")
            
        except Exception as e:
            import traceback
            print(f"Erro ao visualizar dados: {e}")
            print(f"Traceback completo:")
            traceback.print_exc()
            # Fechar popup se houver
            self.close_loading_popup()
    
    def close_loading_popup(self):
        """Fecha o popup de carregamento"""
        if hasattr(self, 'loading_popup') and self.loading_popup:
            self.loading_popup.destroy()
            self.loading_popup = None
    
    def run_update_dates_with_loading(self):
        """Executa a função gerar_update_cards diretamente e abre visualização"""
        self.script_running = True
        
        def simulate_progress():
            """Simula progresso enquanto o script está rodando"""
            progress = 0
            while self.script_running and progress < 0.95:
                progress += 0.10  # Incremento maior para barra mais rápida
                self.root.after(0, lambda p=progress: self.update_progress(p))
                threading.Event().wait(0.1)  # Intervalo menor para resposta mais rápida
        
        def execute():
            try:
                print("Carregando dados de update em memoria...")
                # Chamar a função diretamente (não via subprocess)
                df = update_dates.gerar_update_cards()
                print(f"DataFrame com {len(df)} linhas")
            except Exception as e:
                print(f"Erro ao gerar arquivo: {e}")
                import traceback
                traceback.print_exc()
                df = None
            finally:
                # Parar simulação, completar progresso
                self.script_running = False
                self.root.after(0, lambda: self.update_progress(1.0))
                
                # Abrir visualização passando o DataFrame diretamente
                if df is not None:
                    print(f"Abrindo visualização de dados com DataFrame ({len(df)} linhas)...")
                    self.root.after(200, lambda: self.close_loading_popup())
                    self.root.after(300, lambda: update_dates.abrir_janela_visualizacao(self, df=df))
                else:
                    print("Erro ao gerar dados")
                    self.root.after(200, lambda: self.close_loading_popup())
        
        # Mostrar popup e iniciar execução em thread separada
        self.show_loading_popup("Gerando relatório de datas...")
        
        # Iniciar thread de execução
        exec_thread = threading.Thread(target=execute, daemon=True)
        exec_thread.start()
        
        # Iniciar thread de simulação de progresso
        progress_thread = threading.Thread(target=simulate_progress, daemon=True)
        progress_thread.start()
    
    def run_script_with_loading(self, script_path, script_name, output_file=None, script_args=None):
        """Executa script em thread separada com popup de carregamento"""
        print(f"\n>>> run_script_with_loading INICIADO <<<")
        print(f">>> Script: {script_name}")
        print(f">>> Caminho: {script_path}")
        print(f">>> Argumentos: {script_args}")
        
        self.script_running = True
        script_args = script_args or []
        
        def simulate_progress():
            """Simula progresso enquanto o script está rodando"""
            progress = 0
            while self.script_running and progress < 0.95:
                progress += 0.10  # Incremento maior para barra mais rápida
                self.root.after(0, lambda p=progress: self.update_progress(p))
                threading.Event().wait(0.1)  # Intervalo menor para resposta mais rápida
        
        def execute():
            try:
                print(f"Executando script: {script_path}")
                print(f"Arquivo esperado: {output_file}")
                command = ["python", script_path] + script_args
                print(f"Comando: {' '.join(command)}")
                
                # Usar Popen para capturar saída em tempo real
                process = subprocess.Popen(
                    command,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    encoding='utf-8',
                    errors='replace',
                    bufsize=1,
                    universal_newlines=True
                )
                
                # Ler saída linha por linha e mostrar no log
                for line in iter(process.stdout.readline, ''):
                    if line:
                        line = line.rstrip()
                        print(line)  # Imprimir no console também
                        self.root.after(0, lambda l=line: self.append_log(l))
                
                process.wait()
                
                print(f"Script {script_name} finalizado!")
                print(f"Código de saída: {process.returncode}")
                
            except Exception as e:
                error_msg = f"Erro ao executar o script: {e}"
                print(error_msg)
                self.root.after(0, lambda: self.append_log(error_msg))
            finally:
                success_output_file = output_file

                # Para "Gerar Relatório", localizar o arquivo mais recente em src/temp/jira_cards
                if script_name == "Gerar Relatório":
                    report_dir = os.path.join("src", "temp", "jira_cards")
                    try:
                        os.makedirs(report_dir, exist_ok=True)
                        if os.path.isdir(report_dir):
                            candidates = [
                                os.path.join(report_dir, name)
                                for name in os.listdir(report_dir)
                                if name.startswith("jira_cards ") and name.endswith(".xlsx")
                            ]
                            if candidates:
                                success_output_file = max(candidates, key=os.path.getmtime)
                                print(f"Relatório mais recente encontrado: {success_output_file}")
                    except Exception as e:
                        print(f"Erro ao localizar relatório gerado: {e}")

                # Parar simulação, completar progresso
                self.script_running = False
                self.root.after(0, lambda: self.update_progress(1.0))
                
                # Para "Imprimir OS", manter o popup aberto com botão fechar
                if script_name == "Imprimir OS":
                    self.root.after(500, lambda: self.add_close_button_to_popup())
                    return
                
                # Se for "Adicionar Datas", abrir diretamente a visualização
                if script_name == "Adicionar Datas" and output_file:
                    # Verificar se o arquivo foi gerado
                    print(f"Verificando se arquivo existe: {output_file}")
                    print(f"Arquivo existe: {os.path.exists(output_file)}")
                    print(f"Caminho absoluto: {os.path.abspath(output_file)}")
                    
                    # Tentar caminhos alternativos
                    possible_paths = [
                        output_file,
                        os.path.abspath(output_file),
                        os.path.join(os.getcwd(), "src", "data_update", "update_cards.xlsx"),
                        ".\\src\\data_update\\update_cards.xlsx"
                    ]
                    
                    found_file = None
                    for path in possible_paths:
                        if os.path.exists(path):
                            found_file = path
                            print(f"Arquivo encontrado em: {path}")
                            break
                    
                    if found_file:
                        print("Abrindo visualização de dados...")
                        self.root.after(200, lambda: self.close_loading_popup())
                        try:
                            self.root.after(300, lambda: update_dates.abrir_janela_visualizacao(self))
                        except Exception as e:
                            print(f"Erro ao abrir janela de visualização: {e}")
                            import traceback
                            traceback.print_exc()
                    else:
                        print(f"Arquivo não encontrado em nenhum dos caminhos tentados")
                        self.root.after(200, lambda: self.show_success_message(script_name, output_file))
                else:
                    # Para outros scripts, mostrar mensagem de sucesso
                    self.root.after(200, lambda: self.show_success_message(script_name, success_output_file))
        
        # Mostrar popup e iniciar execução em thread separada
        self.show_loading_popup(f"Executando {script_name}...")
        
        # Iniciar thread de execução
        exec_thread = threading.Thread(target=execute, daemon=True)
        exec_thread.start()
        
        # Iniciar thread de simulação de progresso
        progress_thread = threading.Thread(target=simulate_progress, daemon=True)
        progress_thread.start()

    def parse_card_ids_input(self, raw_text):
        """Converte texto livre em lista de IDs do card, aceitando separadores comuns."""
        if not raw_text:
            return []
        return [value.strip() for value in re.split(r"[,;\s]+", raw_text) if value.strip()]

    def request_card_ids_to_print(self):
        """Solicita os IDs dos cards que devem ser processados no script de impressão."""
        popup = ctk.CTkToplevel(self.root)
        popup.title("Imprimir OS")
        popup.geometry("420x360")
        popup.resizable(False, False)
        popup.transient(self.root)
        popup.grab_set()

        self.root.update_idletasks()
        root_x = self.root.winfo_x()
        root_y = self.root.winfo_y()
        root_width = self.root.winfo_width()
        root_height = self.root.winfo_height()
        popup_width = 420
        popup_height = 360
        x = root_x + (root_width - popup_width) // 2
        y = root_y + (root_height - popup_height) // 2
        popup.geometry(f"{popup_width}x{popup_height}+{x}+{y}")

        label = ctk.CTkLabel(
            popup,
            text="Cole os IDs dos cards (uma por linha, ou separados por vírgula):",
            font=ctk.CTkFont(size=13, weight="bold")
        )
        label.pack(anchor="w", padx=16, pady=(16, 8))

        textbox = ctk.CTkTextbox(popup, height=220)
        textbox.pack(fill="both", expand=True, padx=16)
        textbox.focus_set()

        buttons_frame = ctk.CTkFrame(popup, fg_color="transparent")
        buttons_frame.pack(fill="x", padx=16, pady=14)

        result = {"card_ids": None}

        def on_confirm():
            raw_text = textbox.get("1.0", "end").strip()
            card_ids = self.parse_card_ids_input(raw_text)
            if card_ids:
                print("IDs de card informados para impressão:")
                for card_id in card_ids:
                    print(card_id)
                result["card_ids"] = card_ids
            popup.destroy()

        def on_cancel():
            result["card_ids"] = None
            popup.destroy()

        cancel_btn = ctk.CTkButton(
            buttons_frame,
            text="Cancelar",
            command=on_cancel,
            width=110,
            fg_color="gray40",
            hover_color="gray50"
        )
        cancel_btn.pack(side="right")

        confirm_btn = ctk.CTkButton(
            buttons_frame,
            text="Confirmar",
            command=on_confirm,
            width=110
        )
        confirm_btn.pack(side="right", padx=(0, 8))

        popup.protocol("WM_DELETE_WINDOW", on_cancel)
        self.root.wait_window(popup)
        return result["card_ids"]
    
    def pcp_routine_action(self, routine_name):
        """Ação executada ao clicar em uma rotina PCP"""
        print(f"\n>>> pcp_routine_action chamada <<<")
        print(f">>> Rotina selecionada: '{routine_name}'")
        print(f">>> Tipo: {type(routine_name)}")
        print(f">>> Comparação com 'Imprimir OS': {routine_name == 'Imprimir OS'}")
        
        if routine_name == "Gerar Relatório":
            # Executar o script new_archive.py com popup de carregamento
            script_path = os.path.join("scripts", "new_archive.py")
            output_file = os.path.join("src", "temp", "jira_cards")
            os.makedirs(output_file, exist_ok=True)
            self.run_script_with_loading(script_path, "Gerar Relatório", output_file)
        
        elif routine_name == "Adicionar Datas":
            # Executar função diretamente ao invés de subprocess
            self.run_update_dates_with_loading()

        elif routine_name == "Imprimir OS":
            print("\n" + "="*60)
            print(">>> BOTÃO IMPRIMIR OS CLICADO <<<")
            print("="*60)
            # Solicitar IDs do card e executar script de download somente para os itens informados
            card_ids_to_print = self.request_card_ids_to_print()
            print(f"IDs retornados do popup: {card_ids_to_print}")
            
            if not card_ids_to_print:
                print("Execução cancelada: nenhum ID informado.")
                return

            script_path = os.path.join("scripts", "download_ops", "download_ops")
            print(f"Caminho do script: {script_path}")
            print(f"Script existe? {os.path.exists(script_path)}")
            
            self.run_script_with_loading(
                script_path,
                "Imprimir OS",
                script_args=["--ids", ",".join(card_ids_to_print)]
            )
        
        else:
            print(f"\n⚠️ AVISO: Rotina '{routine_name}' não tem ação definida!")
            print("Rotinas disponíveis:")
            print("  - Gerar Relatório")
            print("  - Adicionar Datas")
            print("  - Imprimir OS")
        
        # Adicione aqui a lógica específica para outras rotinas
    
    def settings_action(self):
        self.set_active_menu_button("settings")
        self.show_content(
            "Configurações",
            "Configure as preferências do sistema de acordo com suas necessidades.",
            title_anchor="w",
            title_img=self.images.get("settings")
        )
    
    def exit_action(self):
        """Fecha a aplicação"""
        self.root.quit()

def main():
    root = ctk.CTk()
    app = SidebarApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
