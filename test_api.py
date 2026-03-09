import requests
import pandas as pd
from requests.auth import HTTPBasicAuth
from datetime import datetime
from dotenv import load_dotenv
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

load_dotenv()

JIRA_URL = os.getenv("JIRA_URL")
EMAIL = os.getenv("EMAIL")
API_TOKEN = os.getenv("API_TOKEN")

# FILTRO JQL
jql = 'project IN (MANTA, TENSYLON) AND status IN ("A Produzir", "Liberado Engenharia")'

url = f"{JIRA_URL}/rest/api/3/search/jql"

headers = {
    "Accept": "application/json"
}

next_page = None
all_rows = []
all_links = []  # Para guardar os links

while True:

    params = {
        "jql": jql,
        "maxResults": 100, 
        "fields": [
            "issuetype",
            "summary",
            "status",
            "customfield_10039",   # SITUAÇÃO
            "customfield_11298",   # VEICULO
            "customfield_10245"    # DT PREVISÃO ENTREGA
        ]
    }

    if next_page:
        params["nextPageToken"] = next_page

    response = requests.get(
        url,
        headers=headers,
        params=params,
        auth=HTTPBasicAuth(EMAIL, API_TOKEN)
    )

    data = response.json()
    issues = data.get("issues", [])

    for issue in issues:

        fields = issue.get("fields", {})

        # Tipo
        tipo = fields.get("issuetype", {}).get("name", "")

        # Chave e link
        key = issue.get("key", "")
        link = f"{JIRA_URL}/browse/{key}"

        # Campos padrão
        resumo = fields.get("summary", "")
        status = fields.get("status", {}).get("name", "")

        # SITUAÇÃO (dropdown Jira)
        situacao_raw = fields.get("customfield_10039")
        print(f"DEBUG - Issue {key} - situacao_raw: {situacao_raw}, type: {type(situacao_raw)}")
        if isinstance(situacao_raw, dict):
            situacao = situacao_raw.get("value", "")
        else:
            situacao = situacao_raw or ""
        print(f"DEBUG - Issue {key} - situacao final: {situacao}")

        # Filtrar apenas situações desejadas
        situacoes_validas = ["⚪️RECEBIDO ENCAMINHADO", "🟢RECEBIDO LIBERADO"]
        if situacao not in situacoes_validas:
            continue

        # VEICULO (dropdown ou texto)
        veiculo_raw = fields.get("customfield_11298")
        if isinstance(veiculo_raw, dict):
            veiculo = veiculo_raw.get("value", "")
        else:
            veiculo = veiculo_raw or ""

        # DATA PREVISÃO
        previsao_raw = fields.get("customfield_10245", "")
        if previsao_raw:
            try:
                previsao = datetime.strptime(previsao_raw, "%Y-%m-%d").strftime("%d/%m/%Y")
            except:
                previsao = previsao_raw
        else:
            previsao = ""

        all_rows.append({
            "Tipo de issue": tipo,
            "Chave": link,
            "Resumo": resumo,
            "Status": status,
            "SITUAÇÃO": situacao,
            "Veiculo - Marca/Modelo": veiculo,
            "DT. PREVISÃO ENTREGA": previsao
        })
        all_links.append(link)

    print("Cartões coletados:", len(all_rows))

    if data.get("isLast"):
        break

    next_page = data.get("nextPageToken")

print("Total de cartões:", len(all_rows))

# Criar dataframe
df = pd.DataFrame(all_rows)

# Gerar Excel
filename = ".\\archives\\jira_cards.xlsx"
df.to_excel(filename, index=False)

# Adicionar hyperlinks na coluna Chave
wb = load_workbook(filename)
ws = wb.active

# Percorrer as linhas e adicionar hyperlinks (começando da linha 2, pulando o cabeçalho)
for idx, link in enumerate(all_links, start=2):
    cell = ws[f'B{idx}']  # Coluna B é a coluna "Chave"
    cell.hyperlink = link
    cell.style = "Hyperlink"

wb.save(filename)

print(f"Arquivo {filename} criado com sucesso!")


# Filtro para identificar cartões sem previsão de entrega preenchida
# Filtrar linhas onde DT. PREVISÃO ENTREGA está vazia
cards_sem_previsao = df[df["DT. PREVISÃO ENTREGA"].isna() | (df["DT. PREVISÃO ENTREGA"] == "")]

# Selecionar apenas as colunas desejadas
cards_update = cards_sem_previsao[[
    "Tipo de issue",
    "Chave",
    "Resumo",
    "DT. PREVISÃO ENTREGA"
]]

# Salvar em novo arquivo Excel
cards_update.to_excel(".\\archives\\update_cards.xlsx", index=False)

print("Arquivo update_cards.xlsx criado com sucesso!")