import sys
import io

# Configurar saída UTF-8 para Windows (apenas se ainda não foi configurado)
if sys.platform == 'win32':
    try:
        if not isinstance(sys.stdout, io.TextIOWrapper) or sys.stdout.encoding != 'utf-8':
            sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
        if not isinstance(sys.stderr, io.TextIOWrapper) or sys.stderr.encoding != 'utf-8':
            sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    except Exception:
        pass  # Já configurado ou não é necessário

import requests
import pandas as pd
from requests.auth import HTTPBasicAuth
from datetime import datetime
from dotenv import load_dotenv
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

load_dotenv()

JIRA_URL = os.getenv("JIRA_URL")
EMAIL = os.getenv("EMAIL")
API_TOKEN = os.getenv("API_TOKEN")

# FILTRO JQL
jql = '(project = MANTA AND status IN ("A Produzir", "Liberado Engenharia")) OR (project = TENSYLON AND status IN ("A Produzir", "Liberado Engenharia", "Aguardando Acabamento", "Aguardando Autoclave", "Aguardando Corte", "Aguardando montagem"))'

url = f"{JIRA_URL}/rest/api/3/search/jql"

headers = {
    "Accept": "application/json"
}

# Criar diretórios de saída se não existirem
jira_cards_dir = os.path.join("src", "temp", "jira_cards")
relatorios_dir = os.path.join(jira_cards_dir, "relatorios")
logs_dir = os.path.join(jira_cards_dir, "logs")
os.makedirs(relatorios_dir, exist_ok=True)
os.makedirs(logs_dir, exist_ok=True)

# Criar arquivo de log com timestamp
log_timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
log_file_path = os.path.join(logs_dir, f"jira_cards_{log_timestamp}.log")
log_file = open(log_file_path, "w", encoding="utf-8")

# Escrever cabeçalho do log
log_file.write(f"{'='*60}\n")
log_file.write(f"EXPORTAÇÃO DE CARTÕES JIRA - INICIADO\n")
log_file.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
log_file.write(f"{'='*60}\n\n")
log_file.write(f"📊 Filtro JQL aplicado:\n{jql}\n\n")
log_file.write(f"{'='*60}\n\n")

next_page = None
all_rows = []
all_links = []  # Para guardar os links

while True:

    params = {
        "jql": jql,
        "maxResults": 100, 
        "fields": [
            "issuetype",
            "summary",
            "status",
            "customfield_10039",   # SITUAÇÃO
            "customfield_11298",   # VEICULO
            "customfield_10245"    # DT PREVISÃO ENTREGA
        ]
    }

    if next_page:
        params["nextPageToken"] = next_page

    response = requests.get(
        url,
        headers=headers,
        params=params,
        auth=HTTPBasicAuth(EMAIL, API_TOKEN)
    )

    data = response.json()
    issues = data.get("issues", [])

    for issue in issues:

        fields = issue.get("fields", {})

        # Tipo
        tipo = fields.get("issuetype", {}).get("name", "")

        # Chave e link
        key = issue.get("key", "")
        link = f"{JIRA_URL}/browse/{key}"

        # Campos padrão
        resumo = fields.get("summary", "")
        status = fields.get("status", {}).get("name", "")

        # SITUAÇÃO (dropdown Jira)
        situacao_raw = fields.get("customfield_10039")
        if isinstance(situacao_raw, dict):
            situacao = situacao_raw.get("value", "")
        else:
            situacao = situacao_raw or ""

        # Filtrar apenas situações desejadas
        situacoes_validas = [
            "⚪️RECEBIDO ENCAMINHADO",
            "🟢RECEBIDO LIBERADO",
            "⚫Aguardando entrada",
        ]
        if situacao not in situacoes_validas:
            continue

        # VEICULO (dropdown ou texto)
        veiculo_raw = fields.get("customfield_11298")
        if isinstance(veiculo_raw, dict):
            veiculo = veiculo_raw.get("value", "")
        else:
            veiculo = veiculo_raw or ""

        # DATA PREVISÃO
        previsao_raw = fields.get("customfield_10245", "")
        if previsao_raw:
            try:
                previsao = datetime.strptime(previsao_raw, "%Y-%m-%d").strftime("%d/%m/%Y")
            except:
                previsao = previsao_raw
        else:
            previsao = ""

        all_rows.append({
            "ID": key,
            "Tipo de issue": tipo,
            "Chave": link,
            "Resumo": resumo,
            "Status": status,
            "SITUAÇÃO": situacao,
            "Veículo": veiculo,
            "DT. PREVISÃO ENTREGA": previsao
        })
        all_links.append(link)

    num_collected = len(all_rows)
    print("Cartões coletados:", num_collected)
    log_file.write(f"📄 Página processada: {num_collected} cartões coletados\n")

    if data.get("isLast"):
        break

    next_page = data.get("nextPageToken")

total_cards = len(all_rows)
print("Total de cartões:", total_cards)
log_file.write(f"\n✅ Total de cartões coletados: {total_cards}\n\n")
log_file.write(f"{'='*60}\n\n")

# Criar dataframe
df = pd.DataFrame(all_rows)

# Gerar Excel com data e hora no nome do arquivo
# Obs.: no Windows não é permitido ':' em nomes de arquivo, por isso usamos '.' entre hora e minuto.
timestamp = datetime.now().strftime("%d.%m.%Y %H.%M")
filename = os.path.join(relatorios_dir, f"jira_cards {timestamp}.xlsx")
df.to_excel(filename, index=False)
log_file.write(f"📝 Criando arquivo Excel...\n")

# Adicionar hyperlinks na coluna Chave
wb = load_workbook(filename)
ws = wb.active

# Percorrer as linhas e adicionar hyperlinks (começando da linha 2, pulando o cabeçalho)
for idx, link in enumerate(all_links, start=2):
    cell = ws[f'C{idx}']  # Coluna C é a coluna "Chave"
    cell.hyperlink = link
    cell.style = "Hyperlink"

wb.save(filename)
log_file.write(f"🔗 Hyperlinks adicionados na coluna 'Chave'\n")
log_file.write(f"💾 Arquivo salvo: {filename}\n\n")

# Escrever resumo no log
log_file.write(f"{'='*60}\n")
log_file.write(f"CONCLUÍDO - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
log_file.write(f"{'='*60}\n")
log_file.write(f"✅ Total de cartões exportados: {total_cards}\n")
log_file.write(f"📁 Arquivo Excel: {os.path.abspath(filename)}\n")
log_file.write(f"📄 Log salvo em: {os.path.abspath(log_file_path)}\n")
log_file.write(f"{'='*60}\n")
log_file.close()

print(f"Arquivo {filename} criado com sucesso!")
print(f"Log salvo em: {log_file_path}")