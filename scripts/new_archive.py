import sys
import io

# Configurar saída UTF-8 para Windows
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import requests
import pandas as pd
from requests.auth import HTTPBasicAuth
from datetime import datetime
from dotenv import load_dotenv
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

load_dotenv()

JIRA_URL = os.getenv("JIRA_URL")
EMAIL = os.getenv("EMAIL")
API_TOKEN = os.getenv("API_TOKEN")

# FILTRO JQL
jql = '(project = MANTA AND status IN ("A Produzir", "Liberado Engenharia")) OR (project = TENSYLON AND status IN ("A Produzir", "Liberado Engenharia", "Aguardando Acabamento", "Aguardando Autoclave", "Aguardando Corte", "Aguardando montagem"))'

url = f"{JIRA_URL}/rest/api/3/search/jql"

headers = {
    "Accept": "application/json"
}

next_page = None
all_rows = []
all_links = []  # Para guardar os links

while True:

    params = {
        "jql": jql,
        "maxResults": 100, 
        "fields": [
            "issuetype",
            "summary",
            "status",
            "customfield_10039",   # SITUAÇÃO
            "customfield_11298",   # VEICULO
            "customfield_10245"    # DT PREVISÃO ENTREGA
        ]
    }

    if next_page:
        params["nextPageToken"] = next_page

    response = requests.get(
        url,
        headers=headers,
        params=params,
        auth=HTTPBasicAuth(EMAIL, API_TOKEN)
    )

    data = response.json()
    issues = data.get("issues", [])

    for issue in issues:

        fields = issue.get("fields", {})

        # Tipo
        tipo = fields.get("issuetype", {}).get("name", "")

        # Chave e link
        key = issue.get("key", "")
        link = f"{JIRA_URL}/browse/{key}"

        # Campos padrão
        resumo = fields.get("summary", "")
        status = fields.get("status", {}).get("name", "")

        # SITUAÇÃO (dropdown Jira)
        situacao_raw = fields.get("customfield_10039")
        if isinstance(situacao_raw, dict):
            situacao = situacao_raw.get("value", "")
        else:
            situacao = situacao_raw or ""

        # Filtrar apenas situações desejadas
        situacoes_validas = [
            "⚪️RECEBIDO ENCAMINHADO",
            "🟢RECEBIDO LIBERADO",
            "⚫Aguardando entrada",
        ]
        if situacao not in situacoes_validas:
            continue

        # VEICULO (dropdown ou texto)
        veiculo_raw = fields.get("customfield_11298")
        if isinstance(veiculo_raw, dict):
            veiculo = veiculo_raw.get("value", "")
        else:
            veiculo = veiculo_raw or ""

        # DATA PREVISÃO
        previsao_raw = fields.get("customfield_10245", "")
        if previsao_raw:
            try:
                previsao = datetime.strptime(previsao_raw, "%Y-%m-%d").strftime("%d/%m/%Y")
            except:
                previsao = previsao_raw
        else:
            previsao = ""

        all_rows.append({
            "ID": key,
            "Tipo de issue": tipo,
            "Chave": link,
            "Resumo": resumo,
            "Status": status,
            "SITUAÇÃO": situacao,
            "Veículo": veiculo,
            "DT. PREVISÃO ENTREGA": previsao
        })
        all_links.append(link)

    print("Cartões coletados:", len(all_rows))

    if data.get("isLast"):
        break

    next_page = data.get("nextPageToken")

print("Total de cartões:", len(all_rows))

# Criar dataframe
df = pd.DataFrame(all_rows)

# Criar diretório de saída se não existir
output_dir = os.path.join("src", "temp", "jira_cards")
os.makedirs(output_dir, exist_ok=True)

# Gerar Excel com data e hora no nome do arquivo
# Obs.: no Windows não é permitido ':' em nomes de arquivo, por isso usamos '.' entre hora e minuto.
timestamp = datetime.now().strftime("%d.%m.%Y %H.%M")
filename = os.path.join(output_dir, f"jira_cards {timestamp}.xlsx")
df.to_excel(filename, index=False)

# Adicionar hyperlinks na coluna Chave
wb = load_workbook(filename)
ws = wb.active

# Percorrer as linhas e adicionar hyperlinks (começando da linha 2, pulando o cabeçalho)
for idx, link in enumerate(all_links, start=2):
    cell = ws[f'C{idx}']  # Coluna C é a coluna "Chave"
    cell.hyperlink = link
    cell.style = "Hyperlink"

wb.save(filename)

print(f"Arquivo {filename} criado com sucesso!")