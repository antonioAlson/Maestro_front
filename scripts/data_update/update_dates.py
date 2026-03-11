import sys
import io

# Configurar saída UTF-8 para Windows
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import pandas as pd
import requests
from requests.auth import HTTPBasicAuth
from datetime import datetime
from dotenv import load_dotenv
import os
from openpyxl import load_workbook
import customtkinter as ctk

load_dotenv()

# CONFIGURAÇÕES
JIRA_URL = os.getenv("JIRA_URL")
EMAIL = os.getenv("EMAIL")
API_TOKEN = os.getenv("API_TOKEN")

# FILTRO JQL - mesma query do generate_archives
jql = '(project = MANTA AND status IN ("A Produzir", "Liberado Engenharia")) OR (project = TENSYLON AND status IN ("A Produzir", "Liberado Engenharia", "Aguardando Acabamento", "Aguardando Autoclave", "Aguardando Corte", "Aguardando montagem"))'

url = f"{JIRA_URL}/rest/api/3/search/jql"

headers = {
    "Accept": "application/json"
}

auth = HTTPBasicAuth(EMAIL, API_TOKEN)

# ============================================================================
# PARTE 1: GERAR ARQUIVO update_cards.xlsx
# ============================================================================

print("Buscando dados do Jira...")

next_page = None
all_rows = []
all_links = []

while True:
    params = {
        "jql": jql,
        "maxResults": 100,
        "fields": [
            "issuetype",
            "summary",
            "status",
            "customfield_10039",   # SITUAÇÃO
            "customfield_11298",   # VEÍCULO
            "customfield_10245"    # DT PREVISÃO ENTREGA
        ]
    }

    if next_page:
        params["nextPageToken"] = next_page

    response = requests.get(
        url,
        headers=headers,
        params=params,
        auth=auth
    )

    if response.status_code != 200:
        print(f"Erro na requisição: {response.status_code}")
        print(response.text)
        break

    data = response.json()
    issues = data.get("issues", [])

    if not issues:
        break

    for issue in issues:
        fields = issue.get("fields", {})

        # Tipo
        tipo = fields.get("issuetype", {}).get("name", "")

        # Chave e link
        key = issue.get("key", "")
        link = f"{JIRA_URL}/browse/{key}"

        # Campos padrão
        resumo = fields.get("summary", "")
        status = fields.get("status", {}).get("name", "")

        # SITUAÇÃO (dropdown Jira)
        situacao_raw = fields.get("customfield_10039")
        if isinstance(situacao_raw, dict):
            situacao = situacao_raw.get("value", "")
        else:
            situacao = situacao_raw or ""
        
        # Filtrar apenas situações desejadas
        situacoes_validas = ["⚪️RECEBIDO ENCAMINHADO", "🟢RECEBIDO LIBERADO"]
        if situacao not in situacoes_validas:
            continue

        # Abreviar situação
        if situacao == "⚪️RECEBIDO ENCAMINHADO":
            situacao = "⚪️REC. Encaminhado"
        elif situacao == "🟢RECEBIDO LIBERADO":
            situacao = "🟢REC. Liberado"

        # VEÍCULO
        veiculo_raw = fields.get("customfield_11298")
        if isinstance(veiculo_raw, dict):
            veiculo = veiculo_raw.get("value", "")
        else:
            veiculo = veiculo_raw or ""

        # DT PREVISÃO ENTREGA
        dt_previsao_raw = fields.get("customfield_10245", "")
        if dt_previsao_raw:
            try:
                dt_previsao = datetime.strptime(dt_previsao_raw, "%Y-%m-%d").strftime("%d/%m/%Y")
            except:
                dt_previsao = dt_previsao_raw
        else:
            dt_previsao = ""

        all_rows.append({
            "ID": key,
            "Tipo de issue": tipo,
            "Chave": link,
            "Resumo": resumo,
            "Status": status,
            "SITUAÇÃO": situacao,
            "Veículo": veiculo,
            "DT. PREVISÃO ENTREGA": dt_previsao
        })

        all_links.append(link)

    print("Cartões coletados:", len(all_rows))

    if data.get("isLast"):
        break

    next_page = data.get("nextPageToken")

print(f"Total de cartões: {len(all_rows)}")

# Criar DataFrame
df = pd.DataFrame(all_rows)

# Filtrar cartões sem previsão de entrega
cards_sem_previsao = df[df["DT. PREVISÃO ENTREGA"].isna() | (df["DT. PREVISÃO ENTREGA"] == "")]

print(f"Cartões sem previsão de entrega: {len(cards_sem_previsao)}")

# Selecionar apenas as colunas desejadas
cards_update = cards_sem_previsao[[
    "ID",
    "Tipo de issue",
    "Chave",
    "Resumo",
    "Status",
    "SITUAÇÃO",
    "Veículo",
    "DT. PREVISÃO ENTREGA"
]]

# Criar diretórios se não existirem
os.makedirs(".\\src\\data_update", exist_ok=True)

# Salvar arquivo
update_filename = ".\\src\\data_update\\update_cards.xlsx"
cards_update.to_excel(update_filename, index=False)

# Adicionar hyperlinks na coluna Chave
wb_update = load_workbook(update_filename)
ws_update = wb_update.active

# Criar lista de links apenas para cards sem previsão
links_sem_previsao = cards_sem_previsao["Chave"].tolist()

# Adicionar hyperlinks (começando da linha 2, pulando o cabeçalho)
for idx, link in enumerate(links_sem_previsao, start=2):
    cell = ws_update[f'C{idx}']  # Coluna C é a coluna "Chave"
    cell.hyperlink = link
    cell.style = "Hyperlink"

# Ajustar largura da coluna SITUAÇÃO (coluna F) para 80px (aproximadamente 11 unidades)
ws_update.column_dimensions['F'].width = 11

wb_update.save(update_filename)

# ============================================================================
# FUNÇÕES PARA SALVAR E ATUALIZAR JIRA
# ============================================================================

def salvar_e_atualizar_jira(file_path, entries_data, card, excel_window):
    """Salva alterações no Excel e atualiza no Jira"""
    try:
        from datetime import datetime
        
        # Ler arquivo Excel
        df = pd.read_excel(file_path)
        
        # Converter a coluna de data para datetime se necessário
        if "DT. PREVISÃO ENTREGA" in df.columns:
            df["DT. PREVISÃO ENTREGA"] = pd.to_datetime(df["DT. PREVISÃO ENTREGA"], errors='coerce')
        
        # Atualizar todos os valores
        for row_idx, column, entry in entries_data:
            new_value = entry.get().strip()
            
            if column == "DT. PREVISÃO ENTREGA":
                if new_value:
                    try:
                        if "/" in new_value:
                            date_obj = datetime.strptime(new_value, "%d/%m/%Y")
                        elif "-" in new_value and len(new_value) == 10:
                            try:
                                date_obj = datetime.strptime(new_value, "%Y-%m-%d")
                            except:
                                date_obj = datetime.strptime(new_value, "%d-%m-%Y")
                        else:
                            date_obj = pd.to_datetime(new_value)
                        
                        df.at[row_idx, column] = pd.Timestamp(date_obj)
                    except (ValueError, Exception) as e:
                        print(f"Aviso: Formato de data inválido '{new_value}' na linha {row_idx}")
                        df.at[row_idx, column] = pd.NaT
                else:
                    df.at[row_idx, column] = pd.NaT
            else:
                df.at[row_idx, column] = new_value if new_value else ""
        
        # Salvar arquivo
        df.to_excel(file_path, index=False)
        
        # Reaplicar hyperlinks se a coluna "Chave" existir
        if "Chave" in df.columns:
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            ws = wb.active
            
            for idx, link in enumerate(df["Chave"], start=2):
                cell = ws[f'C{idx}']
                cell.hyperlink = link
                cell.style = "Hyperlink"
            
            if 'F' in ws.column_dimensions:
                ws.column_dimensions['F'].width = 11
            
            wb.save(file_path)
        
        print(f"✓ Dados salvos no Excel!")
        
        # Atualizar Jira
        atualizar_jira_dates(file_path, card, excel_window)
        
    except Exception as e:
        print(f"Erro ao salvar alterações: {e}")
        import traceback
        traceback.print_exc()


def atualizar_jira_dates(file_path, card, excel_window):
    """Atualiza as datas no Jira baseado no arquivo Excel"""
    try:
        from datetime import datetime
        
        CAMPO_PREVISAO = "customfield_10245"
        
        # Ler arquivo Excel
        df = pd.read_excel(file_path)
        
        # Filtrar apenas linhas com ID e data preenchidos
        df_update = df[df["ID"].notna() & df["DT. PREVISÃO ENTREGA"].notna()].copy()
        
        if len(df_update) == 0:
            print("Nenhuma data para atualizar no Jira")
            mostrar_resultado_jira(0, 0, card, excel_window)
            return
        
        print(f"Atualizando {len(df_update)} issues no Jira...")
        
        headers = {
            "Accept": "application/json",
            "Content-Type": "application/json"
        }
        
        auth = HTTPBasicAuth(EMAIL, API_TOKEN)
        
        success_count = 0
        error_count = 0
        
        for index, row in df_update.iterrows():
            issue_id = row["ID"]
            data_entrega = row["DT. PREVISÃO ENTREGA"]
            
            # Converter para formato YYYY-MM-DD
            try:
                if isinstance(data_entrega, pd.Timestamp):
                    data_formatada = data_entrega.strftime("%Y-%m-%d")
                elif isinstance(data_entrega, str):
                    if "/" in data_entrega:
                        date_obj = datetime.strptime(data_entrega, "%d/%m/%Y")
                    else:
                        date_obj = datetime.strptime(data_entrega, "%Y-%m-%d")
                    data_formatada = date_obj.strftime("%Y-%m-%d")
                else:
                    print(f"Formato não reconhecido para {issue_id}: {data_entrega}")
                    error_count += 1
                    continue
                
                # Fazer requisição ao Jira
                url = f"{JIRA_URL}/rest/api/3/issue/{issue_id}"
                payload = {
                    "fields": {
                        CAMPO_PREVISAO: data_formatada
                    }
                }
                
                response = requests.put(url, headers=headers, auth=auth, json=payload)
                
                if response.status_code == 204:
                    print(f"✓ {issue_id} atualizado para {data_formatada}")
                    success_count += 1
                else:
                    print(f"✗ Erro ao atualizar {issue_id}: {response.status_code} - {response.text}")
                    error_count += 1
                    
            except Exception as e:
                print(f"✗ Erro ao processar {issue_id}: {e}")
                error_count += 1
        
        # Mostrar resultado
        print(f"\nAtualização concluída: {success_count} sucesso, {error_count} erros")
        mostrar_resultado_jira(success_count, error_count, card, excel_window)
        
    except Exception as e:
        print(f"Erro ao atualizar Jira: {e}")
        import traceback
        traceback.print_exc()


def mostrar_resultado_jira(success_count, error_count, card, excel_window):
    """Mostra resultado da atualização do Jira em um popup"""
    # Criar popup de resultado
    result_popup = ctk.CTkToplevel(excel_window)
    result_popup.title("Resultado da Atualização")
    result_popup.geometry("420x240")
    result_popup.resizable(False, False)
    
    # Centralizar popup em relação à janela de visualização
    result_popup.update_idletasks()
    excel_window.update_idletasks()
    
    width = 420
    height = 240
    
    # Obter posição e tamanho da janela de visualização
    window_x = excel_window.winfo_x()
    window_y = excel_window.winfo_y()
    window_width = excel_window.winfo_width()
    window_height = excel_window.winfo_height()
    
    # Calcular posição centralizada em relação à janela de visualização
    x = window_x + (window_width - width) // 2
    y = window_y + (window_height - height) // 2
    
    result_popup.geometry(f"{width}x{height}+{x}+{y}")
    
    # Trazer para frente
    result_popup.transient(excel_window)
    result_popup.lift()
    result_popup.focus_force()
    result_popup.attributes('-topmost', True)
    result_popup.after(100, lambda: result_popup.attributes('-topmost', False))
    result_popup.grab_set()
    
    # Título
    title = "✓ Atualização Concluída!" if error_count == 0 else "⚠️ Atualização Parcial"
    color = "#90EE90" if error_count == 0 else "orange"
    
    title_label = ctk.CTkLabel(
        result_popup,
        text=title,
        font=ctk.CTkFont(size=20, weight="bold"),
        text_color=color
    )
    title_label.pack(pady=(24, 16))
    
    # Detalhes
    detail_text = f"✓ Sucesso: {success_count}\n✗ Erros: {error_count}"
    detail_label = ctk.CTkLabel(
        result_popup,
        text=detail_text,
        font=ctk.CTkFont(size=16)
    )
    detail_label.pack(pady=(0, 12))
    
    # Mensagem adicional
    if error_count == 0:
        msg = "Todas as datas foram atualizadas com sucesso no Jira!"
    else:
        msg = "Algumas datas apresentaram erro. Verifique o console."
    
    msg_label = ctk.CTkLabel(
        result_popup,
        text=msg,
        font=ctk.CTkFont(size=11),
        text_color="gray70",
        wraplength=350
    )
    msg_label.pack(pady=(0, 18))
    
    # Função para fechar popup e depois a janela
    def fechar_tudo():
        try:
            result_popup.grab_release()
        except Exception:
            pass
        if result_popup.winfo_exists():
            result_popup.destroy()
        if excel_window.winfo_exists():
            excel_window.after(100, excel_window.destroy)

    # Fechar pelo X da janela segue a mesma regra do botão OK
    result_popup.protocol("WM_DELETE_WINDOW", fechar_tudo)
    
    # Botão fechar (fecha popup e janela de visualização)
    close_btn = ctk.CTkButton(
        result_popup,
        text="OK",
        command=fechar_tudo,
        width=180,
        height=50,
        corner_radius=10,
        font=ctk.CTkFont(size=16, weight="bold"),
        fg_color="#2a9d2a" if error_count == 0 else "gray40",
        hover_color="#238a23" if error_count == 0 else "gray50"
    )
    close_btn.pack(pady=(0, 16), padx=20)


def abrir_arquivo_excel(file_path):
    """Abre o arquivo Excel com o aplicativo padrão do sistema"""
    try:
        if os.path.exists(file_path):
            os.startfile(file_path)  # Windows
            print(f"Abrindo arquivo: {file_path}")
        else:
            print(f"Arquivo não encontrado: {file_path}")
    except Exception as e:
        print(f"Erro ao abrir arquivo: {e}")


# ============================================================================
# FUNÇÃO PARA ABRIR JANELA DE VISUALIZAÇÃO
# ============================================================================

def abrir_janela_visualizacao(main_app):
    """Abre janela de visualização dos dados do Excel"""
    file_path = update_filename
    
    try:
        print(f"Iniciando visualização de dados: {file_path}")
        
        # Ler arquivo Excel
        df = pd.read_excel(file_path)
        print(f"Arquivo lido com sucesso. {len(df)} linhas encontradas")
        
        # Criar cópia para visualização
        df_visual = df.copy()
        
        # Lista para armazenar referências aos campos editáveis
        entries_data = []
        
        # Combinar Tipo de issue e Resumo em uma única coluna
        if "Tipo de issue" in df_visual.columns and "Resumo" in df_visual.columns:
            df_visual["Issue"] = df_visual["Tipo de issue"].astype(str) + " - " + df_visual["Resumo"].astype(str)
            df_visual = df_visual.drop(columns=["Tipo de issue", "Resumo"])
        
        # Filtrar colunas (remover ID e Chave)
        columns_to_hide = ["ID", "Chave"]
        display_columns = [col for col in df_visual.columns if col not in columns_to_hide]
        
        # Reorganizar para colocar Issue em primeiro
        if "Issue" in display_columns:
            display_columns.remove("Issue")
            display_columns.insert(0, "Issue")
        
        print(f"Colunas a exibir: {display_columns}")
        print("Criando nova janela...")
        
        # Criar nova janela
        excel_window = ctk.CTkToplevel(main_app.root)
        excel_window.title("Visualização de Dados - " + os.path.basename(file_path))
        excel_window.geometry("900x500")
        
        # Centralizar a nova janela
        excel_window.update_idletasks()
        width = 900
        height = 500
        screen_width = excel_window.winfo_screenwidth()
        screen_height = excel_window.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        excel_window.geometry(f"{width}x{height}+{x}+{y}")
        
        # Trazer janela para frente
        excel_window.lift()
        excel_window.focus_force()
        excel_window.attributes('-topmost', True)
        excel_window.after(100, lambda: excel_window.attributes('-topmost', False))
        
        # Card de conteúdo
        card = ctk.CTkFrame(excel_window, corner_radius=10)
        card.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Título
        title_label = ctk.CTkLabel(
            card,
            text="  Visualização de Dados",
            font=ctk.CTkFont(size=18, weight="bold")
        )
        title_label.pack(pady=10, anchor="w", padx=15)
        
        # Info do arquivo
        info_label = ctk.CTkLabel(
            card,
            text=f"Arquivo: {os.path.basename(file_path)} | Total de registros: {len(df_visual)}",
            font=ctk.CTkFont(size=10),
            text_color="gray70"
        )
        info_label.pack(pady=3, anchor="w", padx=15)
        
        # Frame scrollable para a tabela
        table_frame = ctk.CTkScrollableFrame(card, height=280)
        table_frame.pack(fill="both", expand=True, padx=15, pady=10)
        
        # Criar cabeçalho da tabela
        for col_idx, column in enumerate(display_columns):
            header_frame = ctk.CTkFrame(
                table_frame,
                fg_color="gray25",
                border_width=1,
                border_color="gray40"
            )
            header_frame.grid(row=0, column=col_idx, padx=1, pady=1, sticky="ew")
            
            header = ctk.CTkLabel(
                header_frame,
                text=str(column),
                font=ctk.CTkFont(size=12, weight="bold"),
                anchor="w",
                width=80 if column in ["Issue", "Status"] else 150
            )
            header.pack(padx=8, pady=6, fill="both", expand=True)
        
        # Preencher dados da tabela (limitado a 100 linhas)
        max_rows = min(len(df_visual), 100)
        for row_idx in range(max_rows):
            for col_idx, column in enumerate(display_columns):
                value = df_visual[column].iloc[row_idx]
                cell_text = str(value) if pd.notna(value) else ""
                if len(cell_text) > 50 and column != "DT. PREVISÃO ENTREGA":
                    cell_text = cell_text[:47] + "..."
                
                cell_frame = ctk.CTkFrame(
                    table_frame,
                    fg_color="gray20",
                    border_width=1,
                    border_color="gray30"
                )
                cell_frame.grid(row=row_idx + 1, column=col_idx, padx=1, pady=1, sticky="ew")
                
                if column == "DT. PREVISÃO ENTREGA":
                    entry = ctk.CTkEntry(
                        cell_frame,
                        font=ctk.CTkFont(size=11),
                        width=150,
                        height=28
                    )
                    entry.insert(0, cell_text)
                    entry.pack(padx=4, pady=2, fill="both", expand=True)
                    entries_data.append((row_idx, column, entry))
                else:
                    cell = ctk.CTkLabel(
                        cell_frame, 
                        text=cell_text,
                        font=ctk.CTkFont(size=11),
                        anchor="w",
                        width=80 if column in ["Issue", "Status"] else 150
                    )
                    cell.pack(padx=8, pady=4, fill="both", expand=True)
        
        # Aviso se houver mais linhas
        if len(df) > 100:
            warning_label = ctk.CTkLabel(
                card,
                text=f"⚠️ Exibindo apenas as primeiras 100 linhas de {len(df)}",
                font=ctk.CTkFont(size=9),
                text_color="orange"
            )
            warning_label.pack(pady=3, anchor="w", padx=15)
        
        # Container para botões
        buttons_container = ctk.CTkFrame(card, fg_color="transparent")
        buttons_container.pack(pady=8)
        
        # Botão Salvar e Atualizar Jira
        save_btn = ctk.CTkButton(
            buttons_container,
            text="💾 Salvar e Atualizar Jira",
            command=lambda: salvar_e_atualizar_jira(file_path, entries_data, card, excel_window),
            width=170,
            height=30,
            font=ctk.CTkFont(size=11, weight="bold"),
            fg_color="#2a9d2a",
            hover_color="#238a23"
        )
        save_btn.pack(side="left", padx=3)
        
        # Botão para abrir no Excel
        open_excel_btn = ctk.CTkButton(
            buttons_container,
            text="📄 Abrir Excel",
            command=lambda: abrir_arquivo_excel(file_path),
            width=120,
            height=30,
            font=ctk.CTkFont(size=11, weight="bold")
        )
        open_excel_btn.pack(side="left", padx=3)
        
        # Botão Fechar
        close_btn = ctk.CTkButton(
            buttons_container,
            text="✕ Fechar",
            command=excel_window.destroy,
            width=100,
            height=30,
            font=ctk.CTkFont(size=11, weight="bold"),
            fg_color="gray40",
            hover_color="gray50"
        )
        close_btn.pack(side="left", padx=3)
        
        print("Janela de visualização criada com sucesso!")
        
    except Exception as e:
        import traceback
        print(f"Erro ao visualizar dados: {e}")
        traceback.print_exc()


if __name__ == "__main__":
    print("Script executado diretamente - apenas gerando arquivo Excel")